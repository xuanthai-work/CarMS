# -*- coding: utf-8 -*-
"""Parse 'Lịch làm việc Tháng 7.xlsx' -> web/data/seed.json (best-effort)."""
import openpyxl, re, json, unicodedata, os
from datetime import timedelta

# Đường dẫn tương đối theo vị trí script: web/scripts/ -> ../../<excel>, ../data/seed.json
_HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.environ.get("CARMS_XLSX", os.path.join(_HERE, "..", "..", "Lịch làm việc Tháng 7.xlsx"))
OUT = os.path.join(_HERE, "..", "data", "seed.json")

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[wb.sheetnames[0]]

# --- map column -> date (row 3), row -> time (col B / index 2) ---
DATE_ROW = 3
TIME_COL = 2
FIRST_DATA_COL = 3

col_date = {}
for c in range(FIRST_DATA_COL, ws.max_column + 1):
    v = ws.cell(row=DATE_ROW, column=c).value
    if hasattr(v, "date"):
        col_date[c] = v

row_time = {}
for r in range(5, ws.max_row + 1):
    v = ws.cell(row=r, column=TIME_COL).value
    if v is not None and hasattr(v, "strftime"):
        row_time[r] = v.strftime("%H:%M")
    elif isinstance(v, str) and re.match(r"^\d{1,2}:\d{2}", v):
        row_time[r] = v[:5]

# ---------- money parsing ----------
def money_val(main, frac):
    val = int(main) * 1_000_000
    if frac:
        val += int(frac) * (10 ** (6 - len(frac)))
    return val

def parse_money_token(tok):
    tok = tok.lower().replace(" ", "")
    m = re.match(r"^(\d+)tr(\d*)$", tok)
    if m:
        return money_val(m.group(1), m.group(2))
    m = re.match(r"^(\d+)k$", tok)
    if m:
        return int(m.group(1)) * 1000
    return None

MONEY_RE = re.compile(r"(\d+tr\d*|\d+k)", re.IGNORECASE)

def find_money(text):
    """return list of (value, start, end, raw)"""
    out = []
    for m in MONEY_RE.finditer(text):
        v = parse_money_token(m.group(1))
        if v:
            out.append((v, m.start(), m.end(), m.group(1)))
    return out

PLATE_RE = re.compile(r"\b(\d{2}[A-Z]-?\d{3}\.?\d{2,3})\b")
PLATE_RE2 = re.compile(r"\b(\d{2}[A-Z]\d{4,5})\b")
PHONE_RE = re.compile(r"(0\d{2,3}[\s.]?\d{3}[\s.]?\d{3,4})")
SEATS_RE = re.compile(r"(?i)xe\s*0*(\d{1,2})\b")
NIGHTS_RE = re.compile(r"(\d+)\s*N\s*(\d+)\s*Đ", re.IGNORECASE)
COC_RE = re.compile(r"(?i)c[ọo]c\s*(\d+tr\d*|\d+k)")
GIAO_RE = re.compile(r"(?i)\bgiao\b\s+(.+?)\s+(\d+tr\d*|\d+k)")

def norm_plate(s):
    return s.replace("-", "").replace(".", "").upper()

def clean(s):
    return re.sub(r"\s+", " ", s).strip(" -:,.")

# ---------- collect raw bookings ----------
raw_bookings = []
for c, d in col_date.items():
    for r, t in row_time.items():
        cell = ws.cell(row=r, column=c).value
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        raw_bookings.append((d, t, r, c, text))

# ---------- master data registries ----------
vehicles = {}   # plate -> obj
drivers = {}    # name -> obj
customers = {}  # (name,phone) -> obj
partners = {}   # name -> obj

def vid(plate): return "v-" + norm_plate(plate)
def did(name): return "d-" + re.sub(r"[^a-z0-9]+", "-", strip_accents(name).lower()).strip("-")
def cid(n): return "c-%03d" % n
def pid(name): return "p-" + re.sub(r"[^a-z0-9]+", "-", strip_accents(name).lower()).strip("-")

def strip_accents(s):
    return "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn").replace("đ","d").replace("Đ","D")

bookings = []
cust_seq = 0

for d, t, r, c, text in raw_bookings:
    flat = text.replace("\n", " ")
    b = {
        "date": d.strftime("%Y-%m-%d"),
        "time": t,
        "rawText": text,
        "status": "confirmed",
    }

    # seats
    ms = SEATS_RE.search(flat)
    seats = int(ms.group(1)) if ms else None
    b["seats"] = seats

    # plate
    plate = None
    mp = PLATE_RE.search(flat) or PLATE_RE2.search(flat)
    if mp:
        plate = norm_plate(mp.group(1))
    b["plate"] = plate

    # nights / days
    mn = NIGHTS_RE.search(flat)
    if mn:
        n_days = int(mn.group(1))
        b["nights"] = f"{mn.group(1)}N{mn.group(2)}Đ"
        b["days"] = n_days
    else:
        b["nights"] = None
        b["days"] = 1
    b["endDate"] = (d + timedelta(days=max(0, b["days"] - 1))).strftime("%Y-%m-%d")

    # outsourcing (Giao ...)
    mg = GIAO_RE.search(flat)
    is_out = bool(mg)
    partner_name = None
    partner_cost = None
    if mg:
        partner_name = clean(mg.group(1))
        partner_cost = parse_money_token(mg.group(2))
    b["isOutsourced"] = is_out

    # deposit
    deposit = None
    mc = COC_RE.search(flat)
    dep_span = None
    if mc:
        deposit = parse_money_token(mc.group(1))
        dep_span = mc.span(1)
    b["deposit"] = deposit

    # price = first money not equal to deposit span and not the partner cost span
    price = None
    for v, s, e, rawtok in find_money(flat):
        if dep_span and s >= dep_span[0] and e <= dep_span[1]:
            continue
        # skip "per khách" style tokens right before /khách or /người -> still take as fallback
        price = v
        break
    b["price"] = price

    # payment note
    pay = []
    if re.search(r"(?i)\bck\b", flat): pay.append("ck")
    if re.search(r"(?i)\btm\b", flat): pay.append("tm")
    if re.search(r"(?i)đã (nhận|thu)", flat): pay.append("đã thu")
    b["paymentNote"] = ", ".join(pay)

    # header (before first ':') for driver
    header = flat.split(":", 1)[0]
    parts = [clean(p) for p in header.split(" - ") if clean(p)]
    driver_name = None
    if not is_out and len(parts) >= 2:
        cand = parts[1]
        # skip if candidate looks like a plate
        if not (PLATE_RE.match(cand) or PLATE_RE2.match(cand)):
            driver_name = re.sub(r"\s*\(?0\d[\d\s.]+\)?", "", cand).strip()
            driver_name = clean(driver_name)
            if not driver_name:
                driver_name = None
    b["driverName"] = driver_name

    # customer + route (segments after first ':')
    segs = flat.split(":")
    cust_seg = clean(segs[1]) if len(segs) > 1 else ""
    route = clean(":".join(segs[2:])) if len(segs) > 2 else ""
    # phone
    phone = None
    mph = PHONE_RE.search(cust_seg) or PHONE_RE.search(flat)
    if mph:
        phone = re.sub(r"[\s.]", "", mph.group(1))
    b["phone"] = phone
    # customer name = cust_seg minus phone, first chunk
    cust_name = cust_seg
    if mph and mph.group(1) in cust_seg:
        cust_name = cust_seg.replace(mph.group(1), "")
    cust_name = clean(re.split(r"\s+-\s+", cust_name)[0]) if cust_name else ""
    b["customerName"] = cust_name or None
    b["route"] = route or (clean(segs[1]) if len(segs) > 1 else "")

    # ---- register master data ----
    if plate:
        if plate not in vehicles:
            vehicles[plate] = {"id": vid(plate), "plate": plate, "seats": seats}
        elif seats and not vehicles[plate].get("seats"):
            vehicles[plate]["seats"] = seats
        b["vehicleId"] = vehicles[plate]["id"]
    else:
        b["vehicleId"] = None

    if driver_name:
        if driver_name not in drivers:
            drivers[driver_name] = {"id": did(driver_name), "name": driver_name}
        b["driverId"] = drivers[driver_name]["id"]
    else:
        b["driverId"] = None

    if partner_name:
        if partner_name not in partners:
            partners[partner_name] = {"id": pid(partner_name), "name": partner_name}
        b["partnerId"] = partners[partner_name]["id"]
        b["partnerName"] = partner_name
        b["partnerCost"] = partner_cost
    else:
        b["partnerId"] = None
        b["partnerName"] = None
        b["partnerCost"] = None

    if cust_name:
        key = (cust_name, phone or "")
        if key not in customers:
            cust_seq += 1
            customers[key] = {"id": cid(cust_seq), "name": cust_name, "phone": phone}
        b["customerId"] = customers[key]["id"]
    else:
        b["customerId"] = None

    # lượt (đi/về/khứ hồi): xét mệnh đề chứa biển số xe chính
    _low = flat.lower()
    if "lượt" not in _low and "1 chiều" not in _low and "một chiều" not in _low:
        b["leg"] = "round"
    else:
        _clauses = flat.split(";")
        _own = _clauses[0] if _clauses else flat
        if plate:
            _np = norm_plate(plate)
            for _c in _clauses:
                if _np in norm_plate(_c):
                    _own = _c
                    break
        _o = _own.lower()
        _ownVe = ("lượt về" in _o) or ("chiều về" in _o) or ("về -" in _o)
        _ownDi = ("lượt đi" in _o) or ("chiều đi" in _o) or ("1 chiều" in _o) or ("một chiều" in _o)
        if _ownVe:
            b["leg"] = "return"
        elif _ownDi:
            b["leg"] = "outbound"
        elif "lượt về" in _low:
            b["leg"] = "outbound"
        elif "lượt đi" in _low:
            b["leg"] = "return"
        else:
            b["leg"] = "round"

    bookings.append(b)

# ---------- demo maintenance dates for vehicles (clearly demo) ----------
veh_list = list(vehicles.values())
demo_insp = ["2026-07-24", "2026-08-12", "2026-09-05", "2026-07-31", "2026-10-18", "2026-08-28"]
demo_ins  = ["2026-08-15", "2026-07-28", "2026-11-02", "2026-09-20", "2026-07-22", "2026-12-01"]
seat_model = {7: "Xe 7 chỗ", 16: "Ford Transit 16", 29: "Xe 29 chỗ", 35: "Xe 35 chỗ", 45: "Xe 45 chỗ"}
for i, v in enumerate(veh_list):
    v["model"] = seat_model.get(v.get("seats"), "")
    v["status"] = "active"
    v["inspectionDue"] = demo_insp[i % len(demo_insp)]
    v["insuranceDue"] = demo_ins[i % len(demo_ins)]
    v["note"] = "Hạn ĐK/BH là dữ liệu mẫu — cần cập nhật đúng"

for dv in drivers.values():
    dv["phone"] = None
    dv["licenseClass"] = ""
    dv["type"] = "own"
    dv["note"] = ""
for c_ in customers.values():
    c_.setdefault("note", "")
for p_ in partners.values():
    p_["phone"] = None
    p_["note"] = "Đối tác nhận giao xe"

# assign booking ids
for i, b in enumerate(sorted(bookings, key=lambda x: (x["date"], x["time"])), 1):
    b["id"] = "b-%03d" % i

seed = {
    "meta": {
        "source": "Lịch làm việc Tháng 7.xlsx",
        "note": "Dữ liệu parse tự động (best-effort) từ Excel; rawText giữ nguyên bản gốc.",
    },
    "vehicles": veh_list,
    "drivers": list(drivers.values()),
    "customers": list(customers.values()),
    "partners": list(partners.values()),
    "bookings": sorted(bookings, key=lambda x: (x["date"], x["time"])),
}

import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(seed, f, ensure_ascii=False, indent=2)

print("vehicles:", len(veh_list))
print("drivers :", len(drivers))
print("customers:", len(customers))
print("partners:", len(partners))
print("bookings:", len(bookings))
print("OUT ->", OUT)
